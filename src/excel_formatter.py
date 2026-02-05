"""Excel formatting and labeling module.

Produces a .xlsx file matching the exact HR hours summary format:
  - Row 1: date range (start date in A1, end date in B1)
  - Row 2: column headers (bold)
  - Row 3+: data rows with numeric hours formatted to 2 decimal places
  - Two sheet tabs: "Hours Summary Report" and the start date (MM.DD.YYYY)
  - Calculated columns: Paid Hours = Regular Hours + OT1 Hours
"""

import logging
import os
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Column order in the output spreadsheet
OUTPUT_COLUMNS = [
    "Employee Number",
    "Last Name",
    "First Name",
    "PayType Name",
    "Regular Hours",
    "OT1 Hours",
    "Paid Hours",
    "Unpaid Hours",
]

HEADER_FONT = Font(name="Calibri", size=11, bold=True)
DATA_FONT = Font(name="Calibri", size=11)
DATE_FONT = Font(name="Calibri", size=11, bold=True)


def _find_column(headers: List[str], *possible_names: str) -> int | None:
    """Return the index of the first matching header (case-insensitive), or None."""
    lower_headers = [h.strip().lower() for h in headers]
    for name in possible_names:
        if name.lower() in lower_headers:
            return lower_headers.index(name.lower())
    return None


def _to_float(value: str) -> float:
    """Convert a string value to float, returning 0.0 on failure."""
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def _build_row_data(
    row: List[str],
    col_map: dict[str, int | None],
) -> dict[str, object]:
    """Extract and compute the output fields for a single data row."""

    def get(col_name: str, default: str = "") -> str:
        idx = col_map.get(col_name)
        if idx is not None and idx < len(row):
            return row[idx]
        return default

    regular = _to_float(get("Regular Hours"))
    ot1 = _to_float(get("OT1 Hours"))
    unpaid = _to_float(get("Unpaid Hours"))
    paid = regular + ot1  # Calculated field

    return {
        "Employee Number": get("Employee Number"),
        "Last Name": get("Last Name"),
        "First Name": get("First Name"),
        "PayType Name": get("PayType Name"),
        "Regular Hours": regular,
        "OT1 Hours": ot1,
        "Paid Hours": paid,
        "Unpaid Hours": unpaid,
    }


def format_excel(
    headers: List[str],
    rows: List[List[str]],
    output_dir: str,
    filename_prefix: str,
    start_date: str | None = None,
    end_date: str | None = None,
) -> str:
    """Create the hours summary Excel workbook matching the required format.

    Args:
        headers: Column header names from the Google Sheet.
        rows: Data rows from the Google Sheet (list of lists).
        output_dir: Directory to write the output file.
        filename_prefix: Prefix for the output filename.
        start_date: Pay period start date as MM/DD/YYYY. Defaults to today.
        end_date: Pay period end date as MM/DD/YYYY. Defaults to today.

    Returns:
        The full path to the generated .xlsx file.
    """
    now = datetime.now()

    if not start_date:
        start_date = now.strftime("%m/%d/%Y")
    if not end_date:
        end_date = now.strftime("%m/%d/%Y")

    # Parse start date for the tab name (MM.DD.YYYY)
    try:
        parsed_start = datetime.strptime(start_date, "%m/%d/%Y")
        tab_date_label = parsed_start.strftime("%m.%d.%Y")
    except ValueError:
        tab_date_label = start_date.replace("/", ".")

    timestamp = now.strftime("%Y%m%d_%H%M%S")
    filename = f"{filename_prefix}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    os.makedirs(output_dir, exist_ok=True)

    # -- Map source columns to output columns ---------------------------------
    col_map = {
        "Employee Number": _find_column(headers, "Employee Number", "Employee ID", "EmpNo"),
        "Last Name": _find_column(headers, "Last Name", "LastName", "Surname"),
        "First Name": _find_column(headers, "First Name", "FirstName", "Given Name"),
        "PayType Name": _find_column(headers, "PayType Name", "PayType", "Pay Type"),
        "Regular Hours": _find_column(headers, "Regular Hours", "Regular Hrs", "Reg Hours"),
        "OT1 Hours": _find_column(headers, "OT1 Hours", "OT1 Hrs", "Overtime Hours", "OT Hours"),
        "Unpaid Hours": _find_column(headers, "Unpaid Hours", "Unpaid Hrs"),
    }

    # -- Build workbook -------------------------------------------------------
    wb = Workbook()

    # === Sheet 1: "Hours Summary Report" =====================================
    ws = wb.active
    ws.title = "Hours Summary Report"

    # Row 1: Date range
    ws.cell(row=1, column=1, value=start_date).font = DATE_FONT
    ws.cell(row=1, column=2, value=end_date).font = DATE_FONT

    # Row 2: Headers
    for col_idx, header_name in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header_name)
        cell.font = HEADER_FONT

    # Row 3+: Data
    numeric_columns = {"Regular Hours", "OT1 Hours", "Paid Hours", "Unpaid Hours"}
    data_start_row = 3
    for row_offset, row_data in enumerate(rows):
        out = _build_row_data(row_data, col_map)
        row_idx = data_start_row + row_offset
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            value = out[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            if col_name in numeric_columns:
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right")

    # Auto-fit column widths
    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        max_len = len(col_name)
        for row_data in rows:
            out = _build_row_data(row_data, col_map)
            max_len = max(max_len, len(str(out[col_name])))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    # === Sheet 2: date tab (MM.DD.YYYY) ======================================
    ws2 = wb.create_sheet(title=tab_date_label)
    # Mirror the same content on the date tab
    ws2.cell(row=1, column=1, value=start_date).font = DATE_FONT
    ws2.cell(row=1, column=2, value=end_date).font = DATE_FONT

    for col_idx, header_name in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws2.cell(row=2, column=col_idx, value=header_name)
        cell.font = HEADER_FONT

    for row_offset, row_data in enumerate(rows):
        out = _build_row_data(row_data, col_map)
        row_idx = data_start_row + row_offset
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            value = out[col_name]
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            if col_name in numeric_columns:
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right")

    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        max_len = len(col_name)
        for row_data in rows:
            out = _build_row_data(row_data, col_map)
            max_len = max(max_len, len(str(out[col_name])))
        ws2.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    wb.save(filepath)
    logger.info("Excel file written to %s", filepath)
    return filepath
